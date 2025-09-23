import os
import stat
import datetime
import pythoncom
import win32com.client as win32
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
import re
import sys
import argparse

EMAIL_LOG_FILE = "backup_email_log.xlsm"  # Macro-enabled file to log email details
METRICS_FILE = "backup_metrics.xlsx"      # File to log backup metrics

DEFAULT_MAILBOX_NAME = "GMailbox"
DEFAULT_BACKUP_ROOT_DIRECTORY = r"C:\EmailBackups"


def separator(char='=', length=50):
    """Returns a simple line separator."""
    return char * length


def show_heading(text):
    """Prints a heading with separators for clarity."""
    print(separator())
    print(text)
    print(separator())


def show_main_menu():
    """
    Display the main menu options.
    Returns the user's choice as an integer or None if invalid.
    """
    show_heading("Email Backup Menu")
    print("1) Backup emails for yesterday")
    print("2) Backup emails for a specific date or range of dates")
    print("q) Quit")
    choice = input("Please select an option (1, 2, or q): ").strip().lower()

    if choice == '1':
        return 1
    elif choice == '2':
        return 2
    elif choice == 'q':
        return 'q'
    else:
        return None


def show_date_menu():
    """
    Display the sub-menu for backup date selection.
    Returns:
      (option, date_list_or_none)
       - option: 1 for single date, 2 for date range, or None if invalid
       - date_list_or_none: 
            * a single [YYYY-MM-DD] if user chooses single date
            * a list of [YYYY-MM-DD, YYYY-MM-DD, ...] for a date range
            * None if user chooses to return or invalid input
    """
    show_heading("Backup Date Selection")
    print("1) A single date")
    print("2) A range of dates")
    print("b) Back to main menu")
    date_choice = input("Select an option (1, 2, or b): ").strip().lower()

    if date_choice == '1':
        date_input = input("Enter the date (YYYY-MM-DD): ")
        # Validate
        try:
            datetime.datetime.strptime(date_input, '%Y-%m-%d')
            return 1, [date_input]
        except ValueError:
            print("Invalid date format. Please try again.")
            return None, None

    elif date_choice == '2':
        start_date_input = input("Enter the start date (YYYY-MM-DD): ")
        end_date_input = input("Enter the end date (YYYY-MM-DD): ")
        # Validate
        try:
            start_date_obj = datetime.datetime.strptime(start_date_input, '%Y-%m-%d')
            end_date_obj = datetime.datetime.strptime(end_date_input, '%Y-%m-%d')
            if start_date_obj > end_date_obj:
                print("Start date cannot be after end date.")
                return None, None

            # Build a list of date strings in the range
            date_list = []
            current = start_date_obj
            while current <= end_date_obj:
                date_list.append(current.strftime('%Y-%m-%d'))
                current += datetime.timedelta(days=1)
            return 2, date_list

        except ValueError:
            print("One or more dates were invalid. Please try again.")
            return None, None

    elif date_choice == 'b':
        return 'b', None

    else:
        return None, None


def parse_cli_args():
    """Parse command-line arguments for non-interactive use."""
    parser = argparse.ArgumentParser(description="Backup Outlook shared mailbox emails.")
    parser.add_argument("--mode", choices=["interactive", "yesterday", "single", "range"], default="interactive", help="Run mode. Default is interactive menu.")
    parser.add_argument("--date", help="Date for single mode in YYYY-MM-DD format.")
    parser.add_argument("--start-date", help="Start date for range mode in YYYY-MM-DD format.")
    parser.add_argument("--end-date", help="End date for range mode in YYYY-MM-DD format.")
    parser.add_argument("--mailbox-name", default=DEFAULT_MAILBOX_NAME, help=f"Shared mailbox display name. Default: {DEFAULT_MAILBOX_NAME}")
    parser.add_argument("--backup-root", default=DEFAULT_BACKUP_ROOT_DIRECTORY, help=f"Root directory for backups. Default: {DEFAULT_BACKUP_ROOT_DIRECTORY}")
    return parser.parse_args()


def sanitize_filename(filename, max_length=100):
    """
    Remove or replace invalid characters from filenames.
    """
    invalid_chars = '<>:"/\\|?*'
    filename = ''.join(c if c not in invalid_chars else '_' for c in filename)
    filename = ''.join(c for c in filename if c.isprintable() and (c.isalnum() or c in ' ._-'))
    filename = filename.strip()
    if len(filename) > max_length:
        filename = filename[:max_length]
    return filename


def truncate_or_fallback_filename(save_directory, subject, max_path_length=255):
    """
    Generate a unique filename by appending a counter if necessary.
    """
    max_filename_length = max_path_length - len(save_directory) - len('.msg') - len(os.sep)
    if max_filename_length <= 0:
        raise Exception("Save directory path is too long.")

    base_subject = subject
    counter = 0
    while True:
        if counter == 0:
            filename = f"{subject}.msg"
        else:
            filename = f"{base_subject}_{counter}.msg"

        if len(filename) > max_filename_length:
            excess_length = len(filename) - max_filename_length
            base_subject = base_subject[:-excess_length]
            if not base_subject:
                base_subject = 'Email_Subject_Changed'
            if counter == 0:
                filename = f"{base_subject}.msg"
            else:
                filename = f"{base_subject}_{counter}.msg"

        full_path = os.path.join(save_directory, filename)
        if not os.path.exists(full_path):
            return filename
        counter += 1


def save_email(item, save_path):
    """
    Save an email as a .msg file.
    """
    try:
        if hasattr(item, "SaveAs"):
            item.SaveAs(save_path, 3)  # 3 refers to the MSG format
            return True
        return False
    except Exception:
        return False


def get_sender_email(message):
    """
    Retrieve the sender's email address by parsing the raw email headers.
    """
    try:
        # Fetch raw email headers
        headers = message.PropertyAccessor.GetProperty(
            "http://schemas.microsoft.com/mapi/proptag/0x007D001E"
        )
        if headers:
            # Extract the email address from the "From:" field in the headers
            match = re.search(r"From:\s.*<(.+?)>", headers)
            if match:
                return match.group(1)

        # Fallback to "Unknown" if no email address is found
        return "Unknown"
    except Exception as e:
        print(f"Failed to retrieve sender email: {str(e)}")
        return "Unknown"


def log_email_details(backup_date, sender_name, sender_email, subject, file_path):
    """
    Log email details into the macro-enabled Excel file (backup_email_log.xlsm).
    Sets file as read-only again after writing.
    """
    try:
        # 1. Remove the read-only flag if file exists
        if os.path.exists(EMAIL_LOG_FILE):
            os.chmod(EMAIL_LOG_FILE, stat.S_IWRITE)

        # 2. Open or create the workbook
        if os.path.exists(EMAIL_LOG_FILE):
            wb = openpyxl.load_workbook(EMAIL_LOG_FILE, keep_vba=True)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Email Logs"
            ws.append(["Date", "Sender Name", "Sender Email", "Subject"])

        # 3. Ensure the 'Email Logs' sheet exists
        if "Email Logs" in wb.sheetnames:
            ws = wb["Email Logs"]
        else:
            ws = wb.create_sheet(title="Email Logs")
            ws.append(["Date", "Sender Name", "Sender Email", "Subject"])

        # 4. Append new data
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=1, value=backup_date)  # Date
        ws.cell(row=new_row, column=2, value=sender_name)  # Sender Name
        ws.cell(row=new_row, column=3, value=sender_email) # Sender Email
        subject_cell = ws.cell(row=new_row, column=4, value=subject)  # Subject
        subject_cell.hyperlink = file_path
        subject_cell.font = Font(color="0000FF", underline="single")

        # 5. Save changes
        wb.save(EMAIL_LOG_FILE)

        # 6. Re-apply the read-only attribute
        os.chmod(EMAIL_LOG_FILE, stat.S_IREAD)

    except Exception as e:
        print(f"Failed to log email details: {str(e)}")


def log_metrics(backup_date, total_emails, saved_emails, fallback_emails, errors):
    """
    Log backup metrics in a separate Excel file (backup_metrics.xlsx).
    Sets file as read-only again after writing.
    """
    try:
        # 1. Remove the read-only flag if file exists
        if os.path.exists(METRICS_FILE):
            os.chmod(METRICS_FILE, stat.S_IWRITE)

        # 2. Open or create the workbook
        if os.path.exists(METRICS_FILE):
            wb = openpyxl.load_workbook(METRICS_FILE)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Backup Metrics"
            ws.append(["Date", "Total Emails", "Saved Emails", "Fallback Emails", "Errors"])

        # 3. Ensure the 'Backup Metrics' sheet exists
        if "Backup Metrics" in wb.sheetnames:
            ws = wb["Backup Metrics"]
        else:
            ws = wb.create_sheet(title="Backup Metrics")
            ws.append(["Date", "Total Emails", "Saved Emails", "Fallback Emails", "Errors"])

        # 4. Append new metrics row
        ws.append([backup_date, total_emails, saved_emails, fallback_emails, errors])

        # 5. Save changes
        wb.save(METRICS_FILE)

        # 6. Re-apply the read-only attribute
        os.chmod(METRICS_FILE, stat.S_IREAD)

    except Exception as e:
        print(f"Failed to log metrics: {str(e)}")


def backup_shared_mailbox(mailbox_name, backup_root_directory, backup_dates):
    """
    Back up all emails from both Inbox and Sent Items of the specified
    shared mailbox for the given dates, storing them in the same folder.
    If there are zero emails for a date, do not create that date's folder.
    """
    pythoncom.CoInitialize()
    try:
        outlook = win32.Dispatch("Outlook.Application").GetNamespace("MAPI")

        # Locate the shared mailbox
        shared_mailbox = None
        for folder in outlook.Folders:
            if folder.Name.lower() == mailbox_name.lower():
                shared_mailbox = folder
                break

        if not shared_mailbox:
            print(f"Could not find the shared mailbox: {mailbox_name}")
            return

        # Attempt to get Inbox and Sent Items
        try:
            inbox_folder = shared_mailbox.Folders["Inbox"]
        except:
            inbox_folder = None
            print(f"Could not find 'Inbox' in {mailbox_name}. Skipping Inbox...")

        try:
            sent_folder = shared_mailbox.Folders["Sent Items"]
        except:
            sent_folder = None
            print(f"Could not find 'Sent Items' in {mailbox_name}. Skipping Sent Items...")

        # Process each date in the backup list
        for backup_date_str in backup_dates:
            backup_date = datetime.datetime.strptime(backup_date_str, '%Y-%m-%d')

            # We'll combine totals for both Inbox and Sent Items
            total_messages_for_day = 0
            saved_emails_for_day = 0
            fallback_for_day = 0
            errors_for_day = 0

            # We'll only create the folder if we actually have emails
            folder_created = False
            save_directory = os.path.join(
                backup_root_directory,
                backup_date.strftime('%Y'),
                backup_date.strftime('%m-%B'),
                backup_date.strftime('%d-%m-%Y')
            )

            def process_folder(folder_obj):
                """
                Restrict to the date range and save each email in this folder.
                Only create the directory if at least 1 email is found.
                """
                nonlocal folder_created
                nonlocal total_messages_for_day
                nonlocal saved_emails_for_day
                nonlocal fallback_for_day
                nonlocal errors_for_day

                if not folder_obj:
                    return  # If folder doesn't exist, skip

                start_date = backup_date.strftime('%m/%d/%Y 00:00')
                end_date = (backup_date + datetime.timedelta(days=1)).strftime('%m/%d/%Y 00:00')
                restriction = f"[ReceivedTime] >= '{start_date}' AND [ReceivedTime] < '{end_date}'"

                messages = folder_obj.Items.Restrict(restriction).Restrict("[MessageClass] <> 'IPM.Outlook.Recall'")
                total_in_this_folder = len(messages)

                print(separator('-', 60))
                print(f"Folder '{folder_obj.Name}': {total_in_this_folder} emails for {backup_date_str}")

                # Update day-level total
                total_messages_for_day += total_in_this_folder

                if total_in_this_folder > 0 and not folder_created:
                    # Create the folder only when we actually have emails
                    os.makedirs(save_directory, exist_ok=True)
                    folder_created = True

                # Local counters for this folder only
                saved_in_this_folder = 0
                fallback_in_this_folder = 0
                errors_in_this_folder = 0

                for index, message in enumerate(messages, start=1):
                    # Progress indicator
                    sys.stdout.write(f"\rProcessing email {index}/{total_in_this_folder}")
                    sys.stdout.flush()

                    try:
                        subject = sanitize_filename(message.Subject or "No Subject")
                        sender_name = getattr(message, "SenderName", "Unknown")
                        sender_email = get_sender_email(message)
                        filename = truncate_or_fallback_filename(save_directory, subject)
                        full_path = os.path.join(save_directory, filename)

                        if save_email(message, full_path):
                            if filename != f"{subject}.msg":
                                fallback_in_this_folder += 1
                            saved_in_this_folder += 1
                            log_email_details(
                                backup_date.strftime('%Y-%m-%d'),
                                sender_name,
                                sender_email,
                                subject,
                                full_path
                            )
                        else:
                            errors_in_this_folder += 1

                    except Exception as e:
                        print(f"\nFailed to process email: {str(e)}")
                        errors_in_this_folder += 1

                print()  # Move to a new line after the progress indicator

                # Add these folder-level counts to the day-level totals
                saved_emails_for_day += saved_in_this_folder
                fallback_for_day += fallback_in_this_folder
                errors_for_day += errors_in_this_folder

            # 1) Process Inbox
            process_folder(inbox_folder)

            # 2) Process Sent Items
            process_folder(sent_folder)

            # Log combined metrics for this date (even if zero)
            log_metrics(
                backup_date.strftime('%Y-%m-%d'),
                total_messages_for_day,
                saved_emails_for_day,
                fallback_for_day,
                errors_for_day
            )

            print(separator('='))
            print(
                f"Backup for {backup_date.strftime('%Y-%m-%d')} completed. "
                f"Total: {total_messages_for_day}, Saved: {saved_emails_for_day}, "
                f"Errors: {errors_for_day}."
            )

            # If folder was never created, we do not have an empty folder.

        # Cleanup
        inbox_folder = None
        sent_folder = None
        shared_mailbox = None
        outlook = None

    except Exception as e:
        print(f"An error occurred: {str(e)}")
    finally:
        pythoncom.CoUninitialize()


if __name__ == "__main__":
    args = parse_cli_args()
    mailbox_name = args.mailbox_name
    backup_root_directory = args.backup_root

    if args.mode != "interactive":
        if args.mode == "yesterday":
            yesterday = datetime.datetime.now() - datetime.timedelta(days=1)
            backup_dates = [yesterday.strftime('%Y-%m-%d')]
        elif args.mode == "single":
            if not args.date:
                print("Error: --date is required when --mode single.")
                sys.exit(1)
            try:
                datetime.datetime.strptime(args.date, '%Y-%m-%d')
            except ValueError:
                print("Error: Invalid --date format. Use YYYY-MM-DD.")
                sys.exit(1)
            backup_dates = [args.date]
        elif args.mode == "range":
            if not args.start_date or not args.end_date:
                print("Error: --start-date and --end-date are required when --mode range.")
                sys.exit(1)
            try:
                start_date_obj = datetime.datetime.strptime(args.start_date, '%Y-%m-%d')
                end_date_obj = datetime.datetime.strptime(args.end_date, '%Y-%m-%d')
            except ValueError:
                print("Error: Invalid date format for --start-date or --end-date. Use YYYY-MM-DD.")
                sys.exit(1)
            if start_date_obj > end_date_obj:
                print("Error: --start-date cannot be after --end-date.")
                sys.exit(1)
            backup_dates = []
            current = start_date_obj
            while current <= end_date_obj:
                backup_dates.append(current.strftime('%Y-%m-%d'))
                current += datetime.timedelta(days=1)
        else:
            print("Error: Unsupported mode provided.")
            sys.exit(1)

        backup_shared_mailbox(mailbox_name, backup_root_directory, backup_dates)
    else:
        while True:
            main_choice = show_main_menu()

            if main_choice == 1:
                yesterday = datetime.datetime.now() - datetime.timedelta(days=1)
                backup_dates = [yesterday.strftime('%Y-%m-%d')]
                backup_shared_mailbox(mailbox_name, backup_root_directory, backup_dates)
                break

            elif main_choice == 2:
                while True:
                    option, date_list = show_date_menu()

                    if option == 'b':
                        break

                    if option is None:
                        print("Invalid option, please try again.")
                        continue

                    backup_shared_mailbox(mailbox_name, backup_root_directory, date_list)
                    break
                break

            elif main_choice == 'q':
                print("Quitting... Goodbye!")
                break

            else:
                print("Invalid choice. Please try again.")
