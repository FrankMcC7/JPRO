# -*- coding: utf-8 -*-
"""
Approved Funds pipeline (robust rewrite)

What it does:
1) Read CSV from a hard-coded path where row 1 is junk and row 2 has REAL HEADERS.
2) Skip the first row at parse time; use row 2 as headers.
3) Save to .xlsx as a proper Excel Table.
4) Count total rows (each row = one approved fund) and count 'Yes' in 'Alternative Fund?'.
5) Compute last Friday and the corresponding Monday–Friday week window.
6) Build and send Outlook email (from a hard-coded account) with:
   - Subject: "ACTION required by CPO: Weekly Alternative Fund Status Review"
   - Bolded date window and fund counts
   - Embedded HTML table of only rows where Alternative Fund? == Yes
   - Attachment: the .xlsx file saved
   - Ending with "Regards," and your 'formal' Outlook signature
7) Print success to terminal.

Note: Run on a machine with Outlook installed & configured for the sender account.
"""

import os
import sys
import csv
import datetime as dt
import pandas as pd

# Install first if needed:
#   pip install pandas openpyxl pywin32
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import win32com.client as win32


# =========================
# ===== USER SETTINGS =====
# =========================
CSV_PATH = r"D:\path\to\your\approved_funds.csv"       # <-- Hard-code the CSV file path
OUTPUT_DIR = r"D:\path\to\your"                         # <-- Hard-code the output folder
SENDER_EMAIL = "your.alias@company.com"                 # <-- Outlook account to send from
TO_LIST = ["recipient1@company.com", "recipient2@company.com"]  # <-- Hard-code TO recipients
CC_LIST = ["cc1@company.com"]                           # <-- Hard-code CC recipients
SIGNATURE_NAME = "formal"                               # <-- Outlook signature name
CSV_ENCODING = "utf-8-sig"                              # <-- Change if your CSV encoding differs

# Columns to include in the embedded HTML table (Alternative Fund? == Yes only)
TABLE_COLUMNS = [
    "Investment Manager",
    "IM CoPER",
    "Fund Name",
    "Fund GCI",
    "PDO",
    "PMA",
    "Alternative Fund?"
]

# Final email subject (per your instruction)
EMAIL_SUBJECT = "ACTION required by CPO: Weekly Alternative Fund Status Review"


# =========================
# ====== UTILITIES ========
# =========================

def get_last_friday(today=None):
    """Return date of the most recent Friday (including today if Friday)."""
    if today is None:
        today = dt.date.today()
    return today - dt.timedelta(days=(today.weekday() - 4) % 7)

def monday_to_friday_window(last_friday_date):
    """Given a Friday date, return (monday_date, friday_date)."""
    return last_friday_date - dt.timedelta(days=4), last_friday_date

def ddmmyyyy(d):
    return d.strftime("%d%m%Y")

def ordinal_day(d):
    """Return '1st', '2nd', '3rd', ... for a date object."""
    n = d.day
    if 11 <= n % 100 <= 13:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"

def normalize_colname(s: str) -> str:
    """Lowercase, strip, drop non-alphanumerics to match variants like 'Alternative Fund ?'."""
    import re
    return re.sub(r"[^a-z0-9]", "", str(s).strip().lower())

def find_column(df: pd.DataFrame, target_name: str) -> str:
    """Find a column by normalized name; tolerate minor punctuation/spacing differences."""
    target_norm = normalize_colname(target_name)
    mapping = {c: normalize_colname(c) for c in df.columns}
    for orig, norm in mapping.items():
        if norm == target_norm:
            return orig
    # Soft fallback: anything that starts with 'alternativefund'
    for orig, norm in mapping.items():
        if norm.startswith("alternativefund"):
            return orig
    raise KeyError(f"Column '{target_name}' not found. Columns seen: {list(df.columns)}")

def load_signature_html(signature_name):
    """Load Outlook signature HTML; return empty string if not found."""
    sig_dir = os.path.join(os.environ.get("APPDATA", ""), r"Microsoft\Signatures")
    if not os.path.isdir(sig_dir):
        return ""
    sig_htm = os.path.join(sig_dir, f"{signature_name}.htm")
    if not os.path.isfile(sig_htm):
        return ""
    try:
        with open(sig_htm, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()
    except Exception:
        return ""


# =========================
# ====== CORE STEPS =======
# =========================

def read_and_prepare_dataframe(csv_path):
    """
    Robust CSV loader for files where:
      - Row 1 = junk/title (single or few fields)
      - Row 2 = REAL HEADERS
    We skip the first row at parse time and let pandas sniff the delimiter.
    """
    # First attempt: sniff delimiter automatically
    try:
        df = pd.read_csv(
            csv_path,
            skiprows=1,          # skip the junk first row
            header=0,            # use the next row as headers
            sep=None,            # auto-detect delimiter
            engine="python",
            dtype=str,           # keep as strings for consistent cleaning
            keep_default_na=False,
            quoting=csv.QUOTE_MINIMAL,
            encoding=CSV_ENCODING
        )
    except Exception as e1:
        # Fallbacks with explicit delimiters
        df = None
        for sep in [",", ";", "|", "\t"]:
            try:
                df = pd.read_csv(
                    csv_path,
                    skiprows=1,
                    header=0,
                    sep=sep,
                    engine="python",
                    dtype=str,
                    keep_default_na=False,
                    quoting=csv.QUOTE_MINIMAL,
                    encoding=CSV_ENCODING
                )
                break
            except Exception:
                df = None
        if df is None:
            raise RuntimeError(
                f"Failed to parse CSV even after fallbacks. Original error: {e1}"
            )

    # Clean headers and trim string cells
    df.columns = [str(c).strip() for c in df.columns]
    for c in df.columns:
        if df[c].dtype == object:
            df[c] = df[c].map(lambda x: x.strip() if isinstance(x, str) else x)

    # Find the Alternative Fund? column (allow minor name variants)
    alt_col = find_column(df, "Alternative Fund?")

    # Normalize Yes/No for that column
    df[alt_col] = df[alt_col].astype(str).str.strip()
    df[alt_col] = df[alt_col].apply(
        lambda v: "Yes" if v.lower() == "yes" else ("No" if v.lower() == "no" else v)
    )

    return df, alt_col

def save_as_xlsx_table(df, out_path, table_name="ApprovedFundsTable"):
    """Save DataFrame to .xlsx with an Excel Table style."""
    df.to_excel(out_path, index=False, sheet_name="Approved Funds")
    wb = load_workbook(out_path)
    ws = wb.active

    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    tbl = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
    tbl.tableStyleInfo = style
    ws.add_table(tbl)

    wb.save(out_path)
    return out_path

def dataframe_to_html_table(df, columns):
    """Return a clean HTML table string for the email body (only provided columns, if present)."""
    cols = [c for c in columns if c in df.columns]
    html = df[cols].to_html(index=False, border=0, justify="left", classes="alt-funds-table")
    style = """
    <style>
      table.alt-funds-table { border-collapse: collapse; font-family: Segoe UI, Arial, sans-serif; font-size: 11pt; }
      .alt-funds-table th, .alt-funds-table td { border: 1px solid #d0d0d0; padding: 6px 8px; }
      .alt-funds-table th { background: #f2f2f2; text-align: left; }
    </style>
    """
    return style + html

def compose_email_html(monday, friday, total_rows, alt_yes_count, table_html, signature_html):
    """Compose final HTML body with bolded dates/counts and action line."""
    mon_str = f"{ordinal_day(monday)} {monday.strftime('%B')}"
    fri_str = f"{ordinal_day(friday)} {friday.strftime('%B')}"
    total_str = f"<b>{total_rows}</b>"
    alt_str = f"<b>{alt_yes_count}</b>"

    body = f"""
    <div style="font-family: Segoe UI, Arial, sans-serif; font-size: 11pt; color:#222;">
      <p>Hi All,</p>

      <p>Please find attached approved funds from <b>{mon_str}</b> to <b>{fri_str}</b> in EMEA region.
      Total {total_str} Funds got approved out of which {alt_str} Funds are marked as Alternative fund.</p>

      <p><b>ACTION FOR CPO:</b> Please review approved funds to check their Alternative fund (Column BF) status
      from the attached list or below table for quick reference-</p>

      {table_html}

      <p>Regards,</p>
      {signature_html}
    </div>
    """
    return body

def create_and_send_email(sender_email, to_list, cc_list, subject, html_body, attachment_path):
    """Create and send Outlook email using the specified account, with attachment."""
    outlook = win32.Dispatch("Outlook.Application")
    mail = outlook.CreateItem(0)

    # Try to send using the specified account (and on behalf in case of shared mailbox)
    try:
        for acct in outlook.Session.Accounts:
            if str(acct.SmtpAddress).lower() == sender_email.lower():
                mail._oleobj_.Invoke(*(64209, 0, 8, 0, acct))  # SendUsingAccount
                break
        mail.SentOnBehalfOfName = sender_email
    except Exception:
        pass

    mail.Subject = subject
    mail.HTMLBody = html_body
    if to_list:
        mail.To = "; ".join(to_list)
    if cc_list:
        mail.CC = "; ".join(cc_list)
    if attachment_path and os.path.isfile(attachment_path):
        mail.Attachments.Add(Source=attachment_path)

    mail.Send()


# =========================
# ========= MAIN ==========
# =========================

def main():
    # Basic guards
    if not os.path.isfile(CSV_PATH):
        print(f"ERROR: CSV not found: {CSV_PATH}")
        sys.exit(1)
    if not os.path.isdir(OUTPUT_DIR):
        print(f"ERROR: Output directory not found: {OUTPUT_DIR}")
        sys.exit(1)

    # Load data (skip junk row; promote real headers)
    df, alt_col = read_and_prepare_dataframe(CSV_PATH)

    # Counts
    total_rows = len(df)  # each row represents one approved fund
    alt_yes_count = (df[alt_col].astype(str).str.strip().str.lower() == "yes").sum()

    # Dates for filename + email window
    last_fri = get_last_friday()
    mon, fri = monday_to_friday_window(last_fri)

    # Save Excel with table using last Friday in DDMMYYYY
    out_filename = f"Approved Funds_{ddmmyyyy(last_fri)}.xlsx"
    out_path = os.path.join(OUTPUT_DIR, out_filename)
    out_path = save_as_xlsx_table(df, out_path)

    # Filter to Alternative Fund? == Yes for the embedded table
    alt_yes_df = df[df[alt_col].astype(str).str.strip().str.lower() == "yes"].copy()
    # Ensure the displayed header matches exactly "Alternative Fund?" in the email
    if alt_col != "Alternative Fund?":
        alt_yes_df.rename(columns={alt_col: "Alternative Fund?"}, inplace=True)
    table_html = dataframe_to_html_table(alt_yes_df, TABLE_COLUMNS)

    # Build email body + signature
    signature_html = load_signature_html(SIGNATURE_NAME)
    body_html = compose_email_html(mon, fri, total_rows, alt_yes_count, table_html, signature_html)

    # Send email
    create_and_send_email(
        SENDER_EMAIL,
        TO_LIST,
        CC_LIST,
        EMAIL_SUBJECT,
        body_html,
        out_path
    )

    print("Email sent and task completed successfully.")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print("ERROR:", str(e))
        sys.exit(1)